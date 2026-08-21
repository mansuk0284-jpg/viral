#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""백화점 리스트.xlsx → artifacts/compete.json + web/assets/compete-data.js

사용자 설명(2026-08-21):
  - 이 자료는 백화점 리스트이자 **삼성(당사) vs LG(X사) 경쟁력** 자료다.
  - `매장명 + MX` 는 **모바일 매장**이다. LG는 모바일 매장이 없으므로 볼 필요 없다.
  - **경쟁력이 의미 있는 데이터 = 삼성·LG가 모두 입점한 백화점**이다.

시트 구조(실측):
  - 6행: 기간 라벨('22년 (년간)', '25.8월', '26년 (누계)' …)
  - 7행: 각 기간마다 `당사 / X사` 두 열
  - 8행: 그 다음 열이 `경쟁력`
  → 기간 1개 = 3열(당사·X사·경쟁력) 묶음. 13열부터 시작.
  - 29행부터 매장 행. 10열이 매장명, 1열 영업팀, 6열 백화점 체인.

경쟁력 = 당사 매출 / X사 매출. 1.0 이면 대등, 1 초과면 우세.
매출 단위는 백만원이며 **금액은 내보내지 않는다**(사내 실적). 경쟁력 배수만 싣는다.

사용: python scripts/build_compete.py
"""
import io
import json
import os
import re
import sys
from datetime import datetime

import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SRC = os.path.join(ROOT, "data", "백화점 리스트.xlsx")
OUT_J = os.path.join(ROOT, "artifacts", "compete.json")
OUT_W = os.path.join(ROOT, "web", "assets", "compete-data.js")
try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

ROW_PERIOD, ROW_SIDE, ROW_COMP, ROW_FIRST = 6, 7, 8, 29
COL_TEAM, COL_CHAIN, COL_NAME = 1, 6, 10


# 주차(W1, 52W…)는 화면에 과하다. 연간·분기·누계·월만 남긴다.
# 뒤쪽에 21년·20년 블록이 또 있어 같은 라벨이 반복되므로 **처음 나온 것만** 쓴다.
KEEP = re.compile(r"^(\d\d년\s?\(년간\)|\d\d년\s?\(누계\)|\d\d\.[上下]|\d\d년\s?[1-4]Q|"
                  r"\d{1,2}월(\s?\(|$)|\d\d\.\d{1,2}월)")


def periods(ws):
    """기간 라벨과 (당사, X사, 경쟁력) 열 위치를 찾는다."""
    out, seen = [], set()
    for c in range(13, ws.max_column + 1):
        lab = ws.cell(ROW_PERIOD, c).value
        if not lab:
            continue
        lab = str(lab).strip()
        if not KEEP.match(lab) or lab in seen:
            continue
        if ws.cell(ROW_SIDE, c).value != "당사":
            continue
        if ws.cell(ROW_SIDE, c + 1).value != "X사":
            continue
        if ws.cell(ROW_COMP, c + 2).value != "경쟁력":
            continue
        seen.add(lab)
        out.append({"label": lab, "ours": c, "theirs": c + 1, "comp": c + 2})
    return out


# 경쟁력 자료와 대시보드의 매장 표기가 다른 것들.
# 공백만 지워선 51/70 밖에 안 붙는다(실측). 나머지는 이름 자체가 다르다.
ALIAS = {
    "현대무역": "현대무역센터", "현대본점": "현대압구정", "갤러리아본점": "갤러리아압구정",
    "신세계타임스퀘어": "신세계영등포", "롯데스타시티": "롯데건대", "롯데김포공항": "롯데김포",
    "현대킨텍스": "현대일산", "롯데중동": "롯데부천", "신세계사우스시티": "신세계시흥",
    "신세계천안아산": "신세계천안", "갤러리아센터시티": "갤러리아천안", "현대충청": "현대청주",
    "롯데부산": "롯데부산본점", "롯데상인": "롯데안동", "현대디큐브": "현대여의도",
}


def norm(name):
    """'현대서울' → 대시보드 표기로 잇기 위한 정규화 키."""
    s = re.sub(r"\s+", "", str(name))
    s = s.replace("(아)", "").replace("(주)", "")
    s = ALIAS.get(s, s)
    # 대시보드는 '더현대 서울'/'갤러리아 진주'를 쓴다 — 양쪽을 같은 모양으로 만든다
    return s.replace("더현대", "현대").replace("갤러리아", "갤")


def main():
    wb = openpyxl.load_workbook(SRC, data_only=True)
    ws = wb.active
    ps = periods(ws)
    if not ps:
        raise SystemExit("기간 열을 찾지 못했다 — 시트 구조가 바뀌었는지 확인할 것")

    stores, skipped_mx, skipped_solo = [], 0, 0
    for r in range(ROW_FIRST, ws.max_row + 1):
        nm = ws.cell(r, COL_NAME).value
        if not nm:
            continue
        nm = str(nm).strip()
        # 매장명 칸에 숫자만 든 행은 매장이 아니라 구분/합계 행이다.
        # (실제로 '9' 라는 유령 매장이 명부에 끼어 72곳으로 세어졌다)
        if not nm or re.fullmatch(r"[\d.\s]+", nm):
            continue
        # 소속팀이 비면 명부 행이 아니다 — 같은 이유로 걸러낸다
        if not str(ws.cell(r, COL_TEAM).value or "").strip():
            continue
        # 모바일 매장 제외 — LG에는 모바일 매장이 없어 비교 자체가 성립하지 않는다
        if nm.endswith("MX") or "MX" in nm:
            skipped_mx += 1
            continue
        row = {"name": nm, "key": norm(nm),
               "team": str(ws.cell(r, COL_TEAM).value or "").strip(),
               "chain": str(ws.cell(r, COL_CHAIN).value or "").strip(),
               "p": {}, "v": {}}
        both = False
        for p in ps:
            o = ws.cell(r, p["ours"]).value or 0
            t = ws.cell(r, p["theirs"]).value or 0
            try:
                o, t = float(o), float(t)
            except (TypeError, ValueError):
                continue
            # 경쟁력은 **양쪽 모두 매출이 있을 때만** 의미가 있다.
            # 한쪽이 0이면 미입점이거나 집계 전이라 배수가 0 또는 무한이 된다.
            if o > 0 and t > 0:
                row["p"][p["label"]] = round(o / t, 4)
                # 금액과 갭 — 2026-08-21 사용자 지시로 싣는다("금액과 갭을 보여줘").
                # 원래는 사내 실적이라 안 실었다. 화면에서는 hover 로만 띄운다.
                # 단위는 시트 그대로 백만원. 갭 = 당사 - X사(양수면 우리가 앞선다).
                row["v"][p["label"]] = [round(o), round(t), round(o - t)]
                both = True
        if not both:
            skipped_solo += 1
            continue
        stores.append(row)

    # 실데이터가 거의 없는 기간은 버린다.
    # 아직 안 온 달(9~12월)과, 시트 뒤쪽 20·21년 블록의 잔재가 1곳씩만 값을 갖는다.
    have_cnt = {}
    for st in stores:
        for k in st["p"]:
            have_cnt[k] = have_cnt.get(k, 0) + 1
    keep = [p["label"] for p in ps if have_cnt.get(p["label"], 0) >= len(stores) * 0.5]
    for st in stores:
        st["p"] = {k: v for k, v in st["p"].items() if k in keep}
        st["v"] = {k: v for k, v in st.get("v", {}).items() if k in keep}

    data = {
        "built": datetime.now().strftime("%Y-%m-%d"),
        "periods": keep,
        "stores": stores,
        "note": "경쟁력 = 당사/X사 매출 배수. 양사 모두 입점한 매장만. v = [당사, X사, 갭] 백만원(사내 실적 — 화면에서는 hover 로만).",
    }
    json.dump(data, io.open(OUT_J, "w", encoding="utf-8"), ensure_ascii=False, indent=1)
    io.open(OUT_W, "w", encoding="utf-8").write(
        "/* build_compete.py 자동생성 — 수정 금지 */\n"
        "window.COMPETE = " + json.dumps(data, ensure_ascii=False, separators=(",", ":")) + ";\n")

    print(f"기간 {len(ps)}개: {', '.join(data['periods'][:6])}{' …' if len(ps) > 6 else ''}")
    print(f"매장 {len(stores)}곳 (모바일 제외 {skipped_mx} · 한쪽만 입점 제외 {skipped_solo})")
    cur = data["periods"][-1] if keep else None
    have = [s for s in stores if cur in s["p"]]
    if have:
        win = sum(1 for s in have if s["p"][cur] >= 1)
        print(f"최신 기간({cur}) 기준 — 우세 {win} / {len(have)}곳")
    print(f"→ {OUT_J}")
    print(f"→ {OUT_W} ({os.path.getsize(OUT_W)//1024}KB)")


if __name__ == "__main__":
    main()
